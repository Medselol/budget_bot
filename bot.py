import os
import asyncio
import sqlite3
from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook

from aiogram import Bot, Dispatcher, F
from aiogram.types import Message, ReplyKeyboardMarkup, KeyboardButton
from aiogram.filters import Command


TOKEN = os.getenv("BOT_TOKEN")

DB_NAME = "budget.db"
EXCEL_FILE = "operations.xlsx"

USD_RATE = 43.5

BUDGETS = ["Влад", "Валера", "Общий"]

EXPENSE_CATEGORIES = [
    "🛒 Продукты",
    "🍔 Кафе",
    "⛽ Топливо",
    "🚕 Такси",
    "🏠 Дом",
    "🚗 Авто",
    "👕 Одежда",
    "💊 Аптека",
    "🎮 Развлечения",
    "✈️ Путешествия",
    "👶 Ребёнок",
    "📱 Связь",
    "✍️ Другое",
]

INCOME_CATEGORIES = [
    "💼 Зарплата",
    "🚗 Автомойка",
    "💸 Возврат долга",
    "🎁 Подарок",
    "💵 Продажа",
    "🏠 Аренда",
    "✍️ Другое",
]

user_state = {}

main_kb = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="➕ Пополнение"), KeyboardButton(text="➖ Расход")],
        [KeyboardButton(text="🔄 Старт месяца"), KeyboardButton(text="💱 Курс USD")],
        [KeyboardButton(text="💰 Баланс")],
        [KeyboardButton(text="📅 Отчёт за день"), KeyboardButton(text="📆 Отчёт за неделю")],
        [KeyboardButton(text="📊 Отчёт за месяц")],
    ],
    resize_keyboard=True,
)

budget_kb = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="👤 Влад"), KeyboardButton(text="👤 Валера")],
        [KeyboardButton(text="👥 Общий")],
        [KeyboardButton(text="⬅️ Назад")],
    ],
    resize_keyboard=True,
)

expense_kb = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🛒 Продукты"), KeyboardButton(text="🍔 Кафе")],
        [KeyboardButton(text="⛽ Топливо"), KeyboardButton(text="🚕 Такси")],
        [KeyboardButton(text="🏠 Дом"), KeyboardButton(text="🚗 Авто")],
        [KeyboardButton(text="👕 Одежда"), KeyboardButton(text="💊 Аптека")],
        [KeyboardButton(text="🎮 Развлечения"), KeyboardButton(text="✈️ Путешествия")],
        [KeyboardButton(text="👶 Ребёнок"), KeyboardButton(text="📱 Связь")],
        [KeyboardButton(text="✍️ Другое")],
        [KeyboardButton(text="⬅️ Назад")],
    ],
    resize_keyboard=True,
)

income_kb = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="💼 Зарплата"), KeyboardButton(text="🚗 Автомойка")],
        [KeyboardButton(text="💸 Возврат долга"), KeyboardButton(text="🎁 Подарок")],
        [KeyboardButton(text="💵 Продажа"), KeyboardButton(text="🏠 Аренда")],
        [KeyboardButton(text="✍️ Другое")],
        [KeyboardButton(text="⬅️ Назад")],
    ],
    resize_keyboard=True,
)


def init_db():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS operations (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT,
        budget TEXT,
        type TEXT,
        category TEXT,
        amount REAL,
        currency TEXT,
        amount_uah REAL,
        amount_usd REAL,
        usd_rate REAL,
        comment TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT
    )
    """)

    cur.execute("INSERT OR IGNORE INTO settings (key, value) VALUES ('chat_id', '')")

    conn.commit()
    conn.close()


def save_chat_id(chat_id):
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute(
        "INSERT OR REPLACE INTO settings (key, value) VALUES ('chat_id', ?)",
        (str(chat_id),)
    )
    conn.commit()
    conn.close()


def get_chat_id():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()
    cur.execute("SELECT value FROM settings WHERE key='chat_id'")
    row = cur.fetchone()
    conn.close()

    if row and row[0]:
        return int(row[0])

    return None


def init_excel():
    if os.path.exists(EXCEL_FILE):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Операции"

    ws.append([
        "Дата",
        "Бюджет",
        "Тип",
        "Категория",
        "Сумма",
        "Валюта",
        "Сумма грн",
        "Сумма USD",
        "Курс USD",
        "Комментарий",
    ])

    wb.save(EXCEL_FILE)


def add_to_excel(date, budget, op_type, category, amount, currency, amount_uah, amount_usd, usd_rate, comment):
    init_excel()

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Операции"]

    ws.append([
        date,
        budget,
        op_type,
        category,
        amount,
        currency,
        amount_uah,
        amount_usd,
        usd_rate,
        comment,
    ])

    wb.save(EXCEL_FILE)


def parse_amount(text):
    parts = text.replace(",", ".").split()

    amount = float(parts[0])
    currency = "UAH"

    if len(parts) > 1:
        cur = parts[1].lower()

        if cur in ["usd", "$", "дол", "доллар", "долларов"]:
            currency = "USD"
        elif cur in ["uah", "грн", "гривна", "гривен"]:
            currency = "UAH"

    comment = " ".join(parts[2:]) if len(parts) > 2 else ""

    if currency == "USD":
        amount_usd = amount
        amount_uah = amount * USD_RATE
    else:
        amount_uah = amount
        amount_usd = amount / USD_RATE

    return amount, currency, amount_uah, amount_usd, USD_RATE, comment


def add_operation(budget, op_type, category, amount, currency, amount_uah, amount_usd, usd_rate, comment):
    date = datetime.now().strftime("%Y-%m-%d %H:%M")

    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()

    cur.execute("""
    INSERT INTO operations 
    (date, budget, type, category, amount, currency, amount_uah, amount_usd, usd_rate, comment)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        date,
        budget,
        op_type,
        category,
        amount,
        currency,
        amount_uah,
        amount_usd,
        usd_rate,
        comment,
    ))

    conn.commit()
    conn.close()

    add_to_excel(
        date,
        budget,
        op_type,
        category,
        amount,
        currency,
        amount_uah,
        amount_usd,
        usd_rate,
        comment,
    )


def get_balance():
    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()

    result = {}

    for budget in BUDGETS:
        cur.execute("""
        SELECT 
            SUM(CASE WHEN type IN ('Пополнение', 'Старт месяца') THEN amount_usd ELSE 0 END),
            SUM(CASE WHEN type = 'Расход' THEN amount_usd ELSE 0 END)
        FROM operations
        WHERE budget = ?
        """, (budget,))

        income, expense = cur.fetchone()

        income = income or 0
        expense = expense or 0

        result[budget] = {
            "income": income,
            "expense": expense,
            "balance": income - expense,
        }

    conn.close()
    return result


def make_report(period):
    now = datetime.now()

    if period == "day":
        title = "за день"
        date_from = now.strftime("%Y-%m-%d 00:00")
    elif period == "week":
        title = "за неделю"
        date_from = (now - timedelta(days=7)).strftime("%Y-%m-%d %H:%M")
    else:
        title = "за месяц"
        date_from = now.strftime("%Y-%m-01 00:00")

    conn = sqlite3.connect(DB_NAME)
    cur = conn.cursor()

    text = f"📊 Отчёт {title}\n"
    text += f"💱 Курс USD: {USD_RATE} грн\n\n"

    total_income = 0
    total_expense = 0

    for budget in BUDGETS:
        cur.execute("""
        SELECT 
            SUM(CASE WHEN type IN ('Пополнение', 'Старт месяца') THEN amount_usd ELSE 0 END),
            SUM(CASE WHEN type = 'Расход' THEN amount_usd ELSE 0 END)
        FROM operations
        WHERE budget = ? AND date >= ?
        """, (budget, date_from))

        income, expense = cur.fetchone()

        income = income or 0
        expense = expense or 0
        balance = income - expense

        total_income += income
        total_expense += expense

        text += (
            f"🔹 {budget}\n"
            f"➕ Доход: {income:.2f} $\n"
            f"➖ Расход: {expense:.2f} $\n"
            f"💰 Остаток: {balance:.2f} $\n"
        )

        cur.execute("""
        SELECT category, SUM(amount_usd)
        FROM operations
        WHERE budget = ? AND type = 'Расход' AND date >= ?
        GROUP BY category
        ORDER BY SUM(amount_usd) DESC
        LIMIT 5
        """, (budget, date_from))

        categories = cur.fetchall()

        if categories:
            text += "🏆 Топ расходов:\n"
            for category, total in categories:
                text += f"• {category}: {total:.2f} $\n"

        text += "\n"



    conn.close()
    return text


bot = Bot(token=TOKEN)
dp = Dispatcher()
USD_RATE = 43.5

@dp.message(Command("setrate"))
async def set_rate(message: Message):
    global USD_RATE

    try:
        rate = float(message.text.split()[1])
        USD_RATE = rate
        await message.answer(f"✅ Новый курс USD: {USD_RATE}")
    except:
        await message.answer("Пример: /setrate 44.2")

@dp.message(Command("start"))
async def start(message: Message):
    save_chat_id(message.chat.id)
    await message.answer(
        "Привет 👋\n"
        "Бот учёта расходов готов ✅\n\n"
        f"Курс USD: {USD_RATE} грн\n"
        "Все балансы показываются в долларах.",
        reply_markup=main_kb,
    )


@dp.message(F.text.in_(["➕ Пополнение", "➖ Расход", "🔄 Старт месяца"]))
async def choose_action(message: Message):
    action = message.text.replace("➕ ", "").replace("➖ ", "").replace("🔄 ", "")

    user_state[message.from_user.id] = {
        "action": action,
        "budget": None,
        "category": None,
        "waiting_custom_category": False,
    }

    await message.answer("Теперь выбери, для кого:", reply_markup=budget_kb)


@dp.message(F.text.in_(["👤 Влад", "👤 Валера", "👥 Общий"]))
async def choose_budget(message: Message):
    state = user_state.get(message.from_user.id)

    if not state:
        await message.answer("Сначала выбери действие.", reply_markup=main_kb)
        return

    budget = message.text.replace("👤 ", "").replace("👥 ", "")
    state["budget"] = budget

    if state["action"] == "Расход":
        await message.answer("Выбери категорию расхода:", reply_markup=expense_kb)
    elif state["action"] == "Пополнение":
        await message.answer("Выбери источник дохода:", reply_markup=income_kb)
    else:
        state["category"] = "Старт месяца"
        await message.answer(
            "Введи стартовую сумму.\n\n"
            "Примеры:\n"
            "1000 грн\n"
            "100 usd\n"
            "100 $"
        )


@dp.message(F.text.in_(EXPENSE_CATEGORIES + INCOME_CATEGORIES))
async def choose_category(message: Message):
    state = user_state.get(message.from_user.id)

    if not state:
        await message.answer("Сначала выбери действие.", reply_markup=main_kb)
        return

    if message.text == "✍️ Другое":
        state["waiting_custom_category"] = True

        if state["action"] == "Расход":
            await message.answer("Напиши, на что потрачено:")
        else:
            await message.answer("Напиши, откуда пришли деньги:")

        return

    state["category"] = message.text

    await message.answer(
        "Введи сумму.\n\n"
        "Примеры:\n"
        "1000 грн\n"
        "100 usd\n"
        "100 $"
    )


@dp.message(F.text == "💱 Курс USD")
async def show_rate(message: Message):
    await message.answer(f"💱 Курс USD сейчас: {USD_RATE} грн")


@dp.message(F.text == "💰 Баланс")
async def balance(message: Message):
    data = get_balance()

    text = "💰 Баланс в долларах:\n\n"

    total = 0

    for budget, values in data.items():
        text += (
            f"🔹 {budget}\n"
            f"➕ Доход: {values['income']:.2f} $\n"
            f"➖ Расход: {values['expense']:.2f} $\n"
            f"💰 Остаток: {values['balance']:.2f} $\n\n"
        )
        total += values["balance"]

    

    await message.answer(text)


@dp.message(F.text == "📅 Отчёт за день")
async def report_day(message: Message):
    await message.answer(make_report("day"))


@dp.message(F.text == "📆 Отчёт за неделю")
async def report_week(message: Message):
    await message.answer(make_report("week"))


@dp.message(F.text == "📊 Отчёт за месяц")
async def report_month(message: Message):
    await message.answer(make_report("month"))


@dp.message(F.text == "⬅️ Назад")
async def back(message: Message):
    await message.answer("Выбери действие:", reply_markup=main_kb)


@dp.message()
async def handle_text(message: Message):
    state = user_state.get(message.from_user.id)

    if not state:
        await message.answer("Сначала выбери действие.", reply_markup=main_kb)
        return

    if state.get("waiting_custom_category"):
        state["category"] = message.text
        state["waiting_custom_category"] = False

        await message.answer(
            "Теперь введи сумму.\n\n"
            "Примеры:\n"
            "1000 грн\n"
            "100 usd\n"
            "100 $"
        )
        return

    budget = state.get("budget")
    action = state.get("action")
    category = state.get("category")

    if not budget:
        await message.answer("Сначала выбери Влад / Валера / Общий.", reply_markup=budget_kb)
        return

    if not category:
        await message.answer("Сначала выбери категорию.", reply_markup=main_kb)
        return

    try:
        amount, currency, amount_uah, amount_usd, rate, comment = parse_amount(message.text)
    except Exception:
        await message.answer(
            "Введи сумму правильно.\n\n"
            "Примеры:\n"
            "1000 грн\n"
            "100 usd\n"
            "100 $"
        )
        return

    add_operation(
        budget=budget,
        op_type=action,
        category=category,
        amount=amount,
        currency=currency,
        amount_uah=amount_uah,
        amount_usd=amount_usd,
        usd_rate=rate,
        comment=comment,
    )

    await message.answer(
        f"✅ Записано\n\n"
        f"Кому: {budget}\n"
        f"Тип: {action}\n"
        f"Категория: {category}\n"
        f"Сумма: {amount:.2f} {currency}\n"
        f"В гривне: {amount_uah:.2f} грн\n"
        f"В долларах: {amount_usd:.2f} $\n"
        f"Курс USD: {rate}",
        reply_markup=main_kb,
    )

    user_state[message.from_user.id] = {}


async def auto_reports():
    while True:
        now = datetime.now()
        chat_id = get_chat_id()

        if chat_id:
            if now.weekday() == 6 and now.hour == 21 and now.minute == 0:
                await bot.send_message(
                    chat_id,
                    "🤖 Автоотчёт за неделю\n\n" + make_report("week")
                )

            if now.day == 1 and now.hour == 9 and now.minute == 0:
                await bot.send_message(
                    chat_id,
                    "🤖 Автоотчёт за месяц\n\n" + make_report("month")
                )

        await asyncio.sleep(60)


async def main():
    init_db()
    init_excel()
    asyncio.create_task(auto_reports())
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
